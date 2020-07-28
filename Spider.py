import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import pickle
from pyquery import PyQuery as pq
from openpyxl import Workbook, load_workbook

browser = webdriver.Chrome()
wait = WebDriverWait(browser, 10)
url = 'https://taobao.com'


def get_cookie(url):
    browser.get(url)
    time.sleep(30)
    cookies = browser.get_cookies()
    with open('cookie.txt', 'wb') as f:
        pickle.dump(cookies, f)
    print('Get Cookies')


def set_cookie(url):
    browser.get(url)
    browser.delete_all_cookies()
    with open('cookie.txt', 'rb') as f:
        cookies = pickle.load(f)
    for cookie in cookies:
        browser.add_cookie(cookie)
        print(cookie)

def search(url, keyword):
        browser.get(url)
        input = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#q'))
        )
        submit = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '#J_TSearchForm > div.search-button > button'))
        )
        input.send_keys(keyword)
        submit.click()

def parse_page():
    #利用Pyquery 查找找到的页面元素
    html = browser.page_source
    doc = pq(html)
    items = doc('#mainsrp-itemlist .items .item').items()
    product_value = []
    for item in items:
        product = {
            'name': item.find('.title').text(),
            'price': item.find('.price').text(),
            'shop': item.find('.shop').text(),
        }

        product_value.append(product)
    print(product_value)
    return product_value

def download_message(message):

    wb = load_workbook('products.xlsx')
    ws = wb.active
    ws.cell(row=1, column=1, value='name')
    ws.cell(row=1, column=2, value='price')
    ws.cell(row=1, column=3, value='shop')
    for x in range(2, len(message) + 2):
        ws.cell(column=1, row=x, value=message[x - 2]['name'])
    for x in range(2, len(message) + 2):
        ws.cell(column=2, row=x, value=message[x - 2]['price'])
    for x in range(2, len(message) + 2):
        ws.cell(column=3, row=x, value=message[x - 2]['shop'])

    wb.save('products.xlsx')

def turn_page(page):
    try:
        browser.refresh()
        input = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, '#mainsrp-pager > div > div > div > div.form > input'))
        )
        submit = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '#mainsrp-pager > div > div > div > div.form > span.btn.J_Submit'))
        )
        #将页面的专业输入框内容清楚，再输入页码
        input.clear()
        input.send_keys(page)
        submit.click()
        #等到当前页面数字，正好对应转页的数字，否则引发错误，重新转页
        wait.until(EC.text_to_be_present_in_element((By.CSS_SELECTOR, '#mainsrp-pager li.item.active> span'), str(page)))
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, '.m-itemlist .items .item')))
    except TimeoutException:
        turn_page(page)

def main():
    browser.get(url)
    print('请手动登陆用户')
    time.sleep(40)

    print('开始搜索页面')
    search(url, 'Macbook PRO')
    #获取产品信息
    product_message = parse_page()
    #获取2~4页的信息
    for i in range(2, 5):
        #转页
        turn_page(i)
        product_message += parse_page()

    #将产品信息存储到excel中
    download_message(product_message)


if __name__ == '__main__':
    main()

